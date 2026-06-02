"""
Generate one Excel template per topic for teacher input.

Each file contains:
  - Lesson List sheet  : lookup table (lesson title → lesson_id)
  - Questions sheet    : pre-filled existing questions (no id column); Lesson
                         column has a dropdown; lesson_id auto-fills via VLOOKUP
  - Challenge+ sheet   : one row per lesson; same dropdown/VLOOKUP pattern

Run after scripts/parseExcel.js so src/data/staticData.js is up to date.
"""

import json, re, subprocess, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

REPO = Path(__file__).parent.parent
OUT_DIR = REPO / "public" / "topic-templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Load real data from staticData.js via Node ──────────────────────────────

node_script = """
import { LESSONS, QUESTIONS, CHALLENGE_PLUS } from './src/data/staticData.js';
process.stdout.write(JSON.stringify({ lessons: LESSONS, questions: QUESTIONS, challengePlus: CHALLENGE_PLUS }));
"""

result = subprocess.run(
    ["node", "--input-type=module"],
    input=node_script.encode(),
    capture_output=True,
    cwd=str(REPO),
)
if result.returncode != 0:
    print("Error reading staticData.js:", result.stderr.decode(), file=sys.stderr)
    sys.exit(1)

data = json.loads(result.stdout)
all_lessons   = data["lessons"]
all_questions = data["questions"]
challenge_map = {c["lesson_id"]: c for c in data["challengePlus"]}

# ── Group lessons by topic ───────────────────────────────────────────────────

topics: dict[str, list] = {}
for l in all_lessons:
    topics.setdefault(l["topic_name"], []).append(l)

# Sort lessons within each topic by lesson_number
for name in topics:
    topics[name].sort(key=lambda l: (len(l["lesson_number"]), l["lesson_number"]))

def safe_filename(name: str) -> str:
    return re.sub(r'[/\\?*\[\]:,&]', "-", name).strip("-").replace("  ", " ")

EXTRA_Q_ROWS = 30

generated = []

for topic_name, lessons in sorted(topics.items()):
    lesson_ids = {l["lesson_id"] for l in lessons}
    topic_qs   = [q for q in all_questions if q["lesson_id"] in lesson_ids]
    lesson_map = {l["lesson_id"]: l["lesson_title"] for l in lessons}
    n_lessons  = len(lessons)
    n_qs       = len(topic_qs)
    last_q_row = 1 + n_qs + EXTRA_Q_ROWS  # 1 = header row

    wb = Workbook()

    # ── Sheet 1: Lesson List ─────────────────────────────────────────────────
    ws_list = wb.active
    ws_list.title = "Lesson List"
    ws_list.append(["Lesson", "lesson_id"])
    ws_list["A1"].font = Font(bold=True)
    ws_list["B1"].font = Font(bold=True)
    for l in lessons:
        ws_list.append([l["lesson_title"], l["lesson_id"]])
    ws_list.column_dimensions["A"].width = 45
    ws_list.column_dimensions["B"].width = 12

    # ── Sheet 2: Questions ───────────────────────────────────────────────────
    ws_q = wb.create_sheet("Questions")
    ws_q.append(["Lesson", "lesson_id", "Question", "Answer", "Scaffold"])
    for cell in ws_q[1]:
        cell.font = Font(bold=True)

    for i, q in enumerate(topic_qs, start=2):
        ws_q.cell(i, 1, lesson_map.get(q["lesson_id"], ""))
        ws_q.cell(i, 2, f"=IF(A{i}=\"\",\"\",VLOOKUP(A{i},'Lesson List'!$A:$B,2,0))")
        ws_q.cell(i, 3, q["question"])
        ws_q.cell(i, 4, q["answer"])
        ws_q.cell(i, 5, q.get("scaffolded", ""))

    # Blank rows for new questions (VLOOKUP ready)
    for r in range(n_qs + 2, last_q_row + 1):
        ws_q.cell(r, 2, f"=IF(A{r}=\"\",\"\",VLOOKUP(A{r},'Lesson List'!$A:$B,2,0))")

    dv_q = DataValidation(
        type="list",
        formula1=f"'Lesson List'!$A$2:$A${n_lessons + 1}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    dv_q.sqref = f"A2:A{last_q_row}"
    ws_q.add_data_validation(dv_q)

    ws_q.column_dimensions["A"].width = 42
    ws_q.column_dimensions["B"].width = 10
    ws_q.column_dimensions["C"].width = 55
    ws_q.column_dimensions["D"].width = 45
    ws_q.column_dimensions["E"].width = 55

    # ── Sheet 3: Challenge+ ──────────────────────────────────────────────────
    ws_cp = wb.create_sheet("Challenge+")
    ws_cp.append(["Lesson", "lesson_id", "Challenge Question", "Challenge Answer"])
    for cell in ws_cp[1]:
        cell.font = Font(bold=True)

    for i, l in enumerate(lessons, start=2):
        existing = challenge_map.get(l["lesson_id"], {})
        ws_cp.cell(i, 1, l["lesson_title"])
        ws_cp.cell(i, 2, f"=IF(A{i}=\"\",\"\",VLOOKUP(A{i},'Lesson List'!$A:$B,2,0))")
        ws_cp.cell(i, 3, existing.get("question", ""))
        ws_cp.cell(i, 4, existing.get("answer", ""))

    dv_cp = DataValidation(
        type="list",
        formula1=f"'Lesson List'!$A$2:$A${n_lessons + 1}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    dv_cp.sqref = f"A2:A{n_lessons + 20}"
    ws_cp.add_data_validation(dv_cp)

    ws_cp.column_dimensions["A"].width = 42
    ws_cp.column_dimensions["B"].width = 10
    ws_cp.column_dimensions["C"].width = 70
    ws_cp.column_dimensions["D"].width = 70

    fname = f"{safe_filename(topic_name)}-template.xlsx"
    wb.save(OUT_DIR / fname)
    generated.append((topic_name, fname))
    print(f"  {fname}  ({n_qs} questions, {n_lessons} lessons)")

print(f"\nGenerated {len(generated)} topic templates → public/topic-templates/")
